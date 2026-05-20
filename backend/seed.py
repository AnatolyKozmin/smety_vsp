"""Заливает в БД:
   - продукты из «список продуктов» (меню_орги_2026.xlsx)
   - блюда из «граммовки» (меню_орги_2026.xlsx), ссылками на продукты
   - людей из «Кто ест» (Copy of Untitled spreadsheet.xlsx)
"""
import os
import sys
import openpyxl

sys.path.insert(0, os.path.dirname(__file__))

from app.database import Base, engine, SessionLocal
from app import models

ROOT = os.path.join(os.path.dirname(__file__), "..")
XLSX_NEW = os.path.join(ROOT, "меню_орги_2026.xlsx")
XLSX_OLD = os.path.join(ROOT, "Copy of Untitled spreadsheet.xlsx")


def _to_float(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        s = str(v).replace(",", ".")
        try:
            return float(s.split("/")[0].split()[0])
        except Exception:
            return default


def _parse_grams_in_package(value, unit, default=1000.0):
    """Excel хранит «вес упаковки» по-разному:
    «0,9/937г» — для «кг»: 0.9 кг = 900 г.
    «500/1000» — для «кг» муки: 500 г.
    «450 нетто» — для маринованных огурцов: 450 г.
    Возвращает значение в граммах."""
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        n = float(value)
    else:
        s = str(value).replace(",", ".")
        try:
            n = float(s.split("/")[0].split()[0])
        except Exception:
            return default
    u = (unit or "").strip().lower()
    if u in ("кг",) and n <= 10:
        n *= 1000.0  # 0.9 кг = 900 г
    return n


def seed_products(db, wb_new):
    if db.query(models.Product).count() > 0:
        print("Продукты уже есть, пропускаю.")
        return
    ws = wb_new["список продуктов"]
    rows = list(ws.iter_rows(values_only=True))
    # header: Продукт, Ед. измерения, Грамм в 1 упаковке, Рублей за ед. измерения, Срок
    for r in rows[1:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue
        name = str(r[0]).strip()
        unit = (str(r[1]).strip() if r[1] else "") or "г"
        gip = _parse_grams_in_package(r[2], unit, 1000.0)
        price = _to_float(r[3], 0.0)
        term = (str(r[4]).strip() if r[4] else "Долгосрочный")
        db.add(models.Product(
            name=name,
            unit=unit,
            grams_in_package=gip,
            price_per_unit=price,
            storage_term=term,
        ))
    db.commit()
    print(f"Продукты залиты: {db.query(models.Product).count()}")


def seed_dishes(db, wb_new):
    if db.query(models.Dish).count() > 0:
        print("Блюда уже есть, пропускаю.")
        return
    products = {p.name.strip(): p for p in db.query(models.Product).all()}

    def find_product(name):
        n = (name or "").strip()
        if not n:
            return None
        if n in products:
            return products[n]
        # «Лук» vs «Лук ». Попробуем без хвоста пробелов.
        for k, p in products.items():
            if k.strip() == n.strip():
                return p
        # автодобавление
        np = models.Product(name=n, unit="г", grams_in_package=1000, price_per_unit=0, storage_term="Долгосрочный")
        db.add(np)
        db.flush()
        products[n] = np
        return np

    ws = wb_new["граммовки"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue
        dish_name = str(r[0]).strip()
        dish = models.Dish(name=dish_name)
        db.add(dish)
        db.flush()
        # пары (имя продукта, граммовка) идут начиная с колонки B
        # колонки: B,C  D,E  F,G  H,I  J,K  L,M  N,O  P,Q  R,S
        for i in range(1, len(r) - 1, 2):
            prod_name = r[i]
            grams = r[i + 1]
            if not prod_name or grams in (None, ""):
                continue
            p = find_product(str(prod_name))
            if p is None:
                continue
            dish.ingredients.append(models.DishIngredient(
                product_id=p.id,
                grams_per_portion=_to_float(grams, 0),
            ))
    db.commit()
    print(f"Блюда залиты: {db.query(models.Dish).count()}")


def seed_people(db, wb_old):
    if db.query(models.Person).count() > 0:
        print("Люди уже есть, пропускаю.")
        return
    ws = wb_old["Кто ест"]
    rows = list(ws.iter_rows(values_only=True))
    current_role = ""
    for r in rows[1:]:
        if not r:
            continue
        role, name, status, *_ = list(r) + [None] * 3
        if name is None or str(name).strip() == "" or str(name).strip().startswith("ИТОГО"):
            continue
        if role:
            current_role = str(role).strip()
        person = models.Person(
            full_name=str(name).strip(),
            role=current_role,
            present=(str(status).strip().lower() == "был") if status else True,
        )
        db.add(person)
    db.commit()
    print(f"Люди залиты: {db.query(models.Person).count()}")


def main():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if os.path.exists(XLSX_NEW):
            wb_new = openpyxl.load_workbook(XLSX_NEW, data_only=True)
            seed_products(db, wb_new)
            seed_dishes(db, wb_new)
        else:
            print(f"⚠️  {XLSX_NEW} не найден — продукты и блюда не залиты.")

        if os.path.exists(XLSX_OLD):
            wb_old = openpyxl.load_workbook(XLSX_OLD, data_only=True)
            seed_people(db, wb_old)
        else:
            print(f"⚠️  {XLSX_OLD} не найден — люди не залиты.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
